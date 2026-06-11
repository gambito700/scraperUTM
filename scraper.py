import requests
from bs4 import BeautifulSoup
import re, json, os, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "indicadores_previsionales.xlsx")
JSON_FILE = os.path.join(BASE_DIR, "dashboard_data.json")

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

def clean_num(s):
    if not s: return None
    s = s.strip().replace('$', '').replace('%', '').replace('*', '').replace('R.I.', '').strip()
    digits = s.replace('.', '').replace(',', '').replace('-', '').strip()
    if not digits.isdigit():
        try: return float(s)
        except: return None
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    elif s.count('.') >= 2:
        s = s.replace('.', '')
    elif s.count('.') == 1:
        parts = s.split('.')
        if len(parts[1]) == 3 and parts[0].isdigit() and parts[1].isdigit():
            s = s.replace('.', '')
    try: return float(s)
    except: return None

def parse_uf(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'VALOR UF' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if 'Mayo' in label: data['uf_mayo'] = clean_num(val)
                    if 'Abril' in label: data['uf_abril'] = clean_num(val)
    return data

def parse_rentas_topes(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'RENTAS TOPES IMPONIBLES' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if 'AFP' in label: data['tope_afp_90uf'] = clean_num(val)
                    if 'IPS' in label or 'INP' in label: data['tope_ips_60uf'] = clean_num(val)
                    if 'Cesant' in label: data['tope_cesantia_135uf'] = clean_num(val)
    return data

def parse_afp(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'TASA COTIZACI' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 5:
                    label = cells[0].get_text(strip=True).strip()
                    if label in ('Capital', 'Cuprum', 'Habitat', 'PlanVital', 'ProVida', 'Modelo', 'Uno'):
                        data[f'afp_{label.lower()}_trabajador'] = clean_num(cells[1].get_text(strip=True))
                        data[f'afp_{label.lower()}_empleador'] = clean_num(cells[2].get_text(strip=True))
                        data[f'afp_{label.lower()}_total'] = clean_num(cells[3].get_text(strip=True))
                        data[f'afp_{label.lower()}_independiente'] = clean_num(cells[4].get_text(strip=True))
    return data

def parse_afc(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'SEGURO DE CESANT' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 3:
                    label = cells[0].get_text(strip=True)
                    emp = cells[1].get_text(strip=True)
                    trab = cells[2].get_text(strip=True) if len(cells) > 2 else ''
                    if 'Indefinido' in label and '11' not in label and 'Casa' not in label:
                        data['afc_indefinido_empleador'] = clean_num(emp)
                        data['afc_indefinido_trabajador'] = clean_num(trab)
                    elif 'Plazo Fijo' in label or 'Plazo fijo' in label:
                        data['afc_plazofijo_empleador'] = clean_num(emp)
                    elif '11' in label:
                        data['afc_11anios_empleador'] = clean_num(emp)
                    elif 'Casa' in label or 'casa' in label:
                        data['afc_casaparticular_empleador'] = clean_num(emp)
    return data

def parse_apv(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'AHORRO PREVISIONAL VOLUNTARIO' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if 'Mensual' in label: data['apv_tope_mensual'] = clean_num(val)
                    if 'Anual' in label: data['apv_tope_anual'] = clean_num(val)
    return data

def parse_deposito_convenido(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'DEP' in txt and 'CONVENIDO' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if 'Anual' in label: data['deposito_convenido_tope_anual'] = clean_num(val)
    return data

def parse_rentas_minimas(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'RENTAS M' in txt and 'IMONIBLES' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if 'Dependientes' in label and 'Independientes' in label:
                        data['renta_minima_dependientes'] = clean_num(val)
                    elif 'Menores' in label or '18' in label:
                        data['renta_minima_menores65'] = clean_num(val)
                    elif 'Casa Particular' in label or 'casa particular' in label:
                        data['renta_minima_casa_particular'] = clean_num(val)
                    elif 'remuneracionales' in label or 'no remuneracional' in label:
                        data['renta_minima_no_remuneracional'] = clean_num(val)
    return data

def parse_seguro_social_sis(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'SEGURO SOCIAL' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if 'Expectativa' in label or 'Vida' in label:
                        data['seguro_social'] = clean_num(val)
        if 'SEGURO DE INVALIDEZ' in txt or 'SIS' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[1].get_text(strip=True)
                    if 'SIS' in label or 'Tasa' in label:
                        data['sis_tasa'] = clean_num(val) if val and val != 'SIS' else None
                    if data.get('sis_tasa') is None and '1,62' in val:
                        data['sis_tasa'] = clean_num(val)
    return data

def parse_salud(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'SALUD' in txt and ('CCAF' in txt or 'FONASA' in txt):
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    val = cells[0].get_text(strip=True) if len(cells) == 2 else cells[1].get_text(strip=True)
                    val2 = cells[1].get_text(strip=True) if len(cells) > 1 else ''
                    if 'CCAF' in label:
                        data['salud_ccaf'] = clean_num(val2)
                    elif 'FONASA' in label:
                        data['salud_fonasa'] = clean_num(val2)
    return data

def parse_trabajos_pesados(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'TRABAJOS PESADOS' in txt or 'trabajos pesados' in txt.lower():
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 4:
                    label = cells[0].get_text(strip=True)
                    emp = cells[2].get_text(strip=True)
                    trab = cells[3].get_text(strip=True)
                    if 'pesado' in label.lower() and 'menos' not in label.lower():
                        data['trabajo_pesado_empleador'] = clean_num(emp)
                        data['trabajo_pesado_trabajador'] = clean_num(trab)
                    elif 'menos' in label.lower():
                        data['trabajo_menos_pesado_empleador'] = clean_num(emp)
                        data['trabajo_menos_pesado_trabajador'] = clean_num(trab)
    return data

def parse_asignacion_familiar(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'ASIGNACI' in txt and 'FAMILIAR' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if len(cells) >= 3:
                    label = cells[0].get_text(strip=True)
                    monto = cells[1].get_text(strip=True)
                    renta = cells[2].get_text(strip=True)
                    if '1' in label or '(A)' in label:
                        data['asignacion_familiar_tramo1'] = clean_num(monto)
                        data['asignacion_familiar_tramo1_renta'] = renta.strip()
                    elif '2' in label or '(B)' in label:
                        data['asignacion_familiar_tramo2'] = clean_num(monto)
                        data['asignacion_familiar_tramo2_renta'] = renta.strip()
                    elif '3' in label or '(C)' in label:
                        data['asignacion_familiar_tramo3'] = clean_num(monto)
                        data['asignacion_familiar_tramo3_renta'] = renta.strip()
                    elif '4' in label or '(D)' in label:
                        data['asignacion_familiar_tramo4_monto'] = 0
                        data['asignacion_familiar_tramo4_renta'] = renta.strip()
    return data

def parse_utm_previred(soup):
    data = {}
    tables = soup.find_all('table')
    for table in tables:
        txt = table.get_text(" ", strip=True)
        if 'UTM' in txt and 'UTA' in txt:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 3:
                    mes = cells[0].get_text(strip=True)
                    utm_val = cells[1].get_text(strip=True)
                    uta_val = cells[2].get_text(strip=True)
                    for m in ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']:
                        if m in mes:
                            data[f'utm_{m.lower()}'] = clean_num(utm_val)
                            data[f'uta_{m.lower()}'] = clean_num(uta_val)
    return data

def scrape_previred():
    url = "https://www.previred.com/indicadores-previsionales/"
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.encoding = 'utf-8'
    soup = BeautifulSoup(r.text, 'lxml')
    data = {}
    data.update(parse_uf(soup))
    data.update(parse_rentas_topes(soup))
    data.update(parse_afp(soup))
    data.update(parse_afc(soup))
    data.update(parse_apv(soup))
    data.update(parse_deposito_convenido(soup))
    data.update(parse_rentas_minimas(soup))
    data.update(parse_seguro_social_sis(soup))
    data.update(parse_salud(soup))
    data.update(parse_trabajos_pesados(soup))
    data.update(parse_asignacion_familiar(soup))
    data.update(parse_utm_previred(soup))
    data['ultima_actualizacion'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return data

def scrape_sii_utm():
    """Scrapes SII UTM/UTA table for 2026"""
    url = "https://www.sii.cl/valores_y_fechas/utm/utm2026.htm"
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.encoding = 'utf-8'
    soup = BeautifulSoup(r.text, 'lxml')

    months_map = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    months_rev = {v: k for k, v in months_map.items()}

    data = {}
    table = soup.find('table', id='table_export')
    if not table:
        table = soup.find('table', class_='table')
    if not table:
        tables = soup.find_all('table', class_=re.compile('table'))
        table = tables[0] if tables else None

    if table:
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 7:
                mes_text = cells[0].get_text(strip=True)
                utm_val = cells[1].get_text(strip=True)
                uta_val = cells[2].get_text(strip=True)
                ipc_val = cells[3].get_text(strip=True)
                var_mensual = cells[4].get_text(strip=True)
                var_acum = cells[5].get_text(strip=True)
                var_anual = cells[6].get_text(strip=True)

                for m_name, m_num in months_map.items():
                    if m_name in mes_text:
                        key = months_rev[m_num]
                        data[f'sii_utm_{key.lower()}'] = clean_num(utm_val) if utm_val.strip() else None
                        data[f'sii_uta_{key.lower()}'] = clean_num(uta_val) if uta_val.strip() else None
                        data[f'sii_ipc_{key.lower()}'] = clean_num(ipc_val) if ipc_val.strip() else None
                        data[f'sii_var_mensual_{key.lower()}'] = clean_num(var_mensual) if var_mensual.strip() else None
                        data[f'sii_var_acum_{key.lower()}'] = clean_num(var_acum) if var_acum.strip() else None
                        data[f'sii_var_anual_{key.lower()}'] = clean_num(var_anual) if var_anual.strip() else None
                        break
    else:
        # Fallback: parse from raw HTML
        rows = soup.select('table tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            if len(cells) >= 7:
                mes_text = cells[0].get_text(strip=True)
                for m_name, m_num in months_map.items():
                    if m_name in mes_text:
                        key = months_rev[m_num]
                        utm_val = cells[1].get_text(strip=True)
                        uta_val = cells[2].get_text(strip=True)
                        ipc_val = cells[3].get_text(strip=True)
                        var_mensual = cells[4].get_text(strip=True)
                        var_acum = cells[5].get_text(strip=True)
                        var_anual = cells[6].get_text(strip=True)
                        data[f'sii_utm_{key.lower()}'] = clean_num(utm_val) if utm_val.strip() else None
                        data[f'sii_uta_{key.lower()}'] = clean_num(uta_val) if uta_val.strip() else None
                        data[f'sii_ipc_{key.lower()}'] = clean_num(ipc_val) if ipc_val.strip() else None
                        data[f'sii_var_mensual_{key.lower()}'] = clean_num(var_mensual) if var_mensual.strip() else None
                        data[f'sii_var_acum_{key.lower()}'] = clean_num(var_acum) if var_acum.strip() else None
                        data[f'sii_var_anual_{key.lower()}'] = clean_num(var_anual) if var_anual.strip() else None
                        break

    data['sii_fecha_consulta'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    return data

def build_sii_monthly_list(data):
    months_es = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre']
    result = []
    for i, m in enumerate(months_es, 1):
        entry = {
            'mes': m.capitalize(),
            'mes_num': i,
            'utm': data.get(f'sii_utm_{m}'),
            'uta': data.get(f'sii_uta_{m}'),
            'ipc': data.get(f'sii_ipc_{m}'),
            'var_mensual': data.get(f'sii_var_mensual_{m}'),
            'var_acum': data.get(f'sii_var_acum_{m}'),
            'var_anual': data.get(f'sii_var_anual_{m}'),
        }
        result.append(entry)
    return result

def add_charts_sheet(wb, sii_monthly):
    chart_sheet_name = "Graficos"
    if chart_sheet_name in wb.sheetnames:
        ws = wb[chart_sheet_name]
        for i in range(ws.max_row, 0, -1):
            ws.delete_rows(i)
    else:
        ws = wb.create_sheet(chart_sheet_name)

    headers = ["Mes", "UTM", "UTA", "IPC", "Var. Mensual", "Var. Acumulada", "Var. Anual"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, entry in enumerate(sii_monthly, 2):
        ws.cell(row=row_idx, column=1, value=entry['mes'])
        ws.cell(row=row_idx, column=2, value=entry['utm'] if entry['utm'] is not None else 0)
        ws.cell(row=row_idx, column=3, value=entry['uta'] if entry['uta'] is not None else 0)
        ws.cell(row=row_idx, column=4, value=entry['ipc'] if entry['ipc'] is not None else 0)
        ws.cell(row=row_idx, column=5, value=entry['var_mensual'] if entry['var_mensual'] is not None else 0)
        ws.cell(row=row_idx, column=6, value=entry['var_acum'] if entry['var_acum'] is not None else 0)
        ws.cell(row=row_idx, column=7, value=entry['var_anual'] if entry['var_anual'] is not None else 0)

    data_end = len(sii_monthly) + 1

    chart1 = LineChart()
    chart1.title = "UTM / UTA 2026"
    chart1.style = 10
    chart1.y_axis.title = "Valor en $"
    chart1.x_axis.title = "Mes"
    chart1.height = 15
    chart1.width = 25

    cats = Reference(ws, min_col=1, min_row=2, max_row=data_end)
    utm_data = Reference(ws, min_col=2, min_row=1, max_row=data_end)
    uta_data = Reference(ws, min_col=3, min_row=1, max_row=data_end)
    chart1.add_data(utm_data, titles_from_data=True)
    chart1.add_data(uta_data, titles_from_data=True)
    chart1.set_categories(cats)

    s1 = chart1.series[0]
    s1.graphicalProperties.line.width = 25000
    s2 = chart1.series[1]
    s2.graphicalProperties.line.width = 25000

    ws.add_chart(chart1, "A12")

    chart2 = LineChart()
    chart2.title = "Variacion IPC 2026"
    chart2.style = 10
    chart2.y_axis.title = "Porcentaje"
    chart2.x_axis.title = "Mes"
    chart2.height = 15
    chart2.width = 25

    var_mensual_data = Reference(ws, min_col=5, min_row=1, max_row=data_end)
    var_acum_data = Reference(ws, min_col=6, min_row=1, max_row=data_end)
    var_anual_data = Reference(ws, min_col=7, min_row=1, max_row=data_end)
    chart2.add_data(var_mensual_data, titles_from_data=True)
    chart2.add_data(var_acum_data, titles_from_data=True)
    chart2.add_data(var_anual_data, titles_from_data=True)
    chart2.set_categories(cats)

    for s in chart2.series:
        s.graphicalProperties.line.width = 25000

    ws.add_chart(chart2, "A30")


def save_to_excel(previred_data, sii_data, sii_monthly):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_sheet(ws, title, headers, rows_data):
        ws.delete_rows(1, ws.max_row) if ws.max_row > 0 else None
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_idx, row in enumerate(rows_data, 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")
        for col_idx, _ in enumerate(headers, 1):
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[chr(64 + col_idx)].width = max(max_len + 4, 20)

    # Sheet 1: Resumen General
    if "Resumen" in wb.sheetnames:
        ws = wb["Resumen"]
    else:
        ws = wb.create_sheet("Resumen")
    wb.active = ws
    resumen_headers = ["Indicador", "Valor"]
    resumen_rows = []
    def add_row(label, key, data_dict, fmt=None):
        if key in data_dict and data_dict[key] is not None:
            val = data_dict[key]
            if isinstance(val, float):
                val = f"{val:,.2f}".replace(',', '@').replace('.', ',').replace('@', '.') if fmt != '%' else f"{val:.2f}%"
            resumen_rows.append([label, str(val)])

    add_row("UF - 31 Mayo 2026", "uf_mayo", previred_data)
    add_row("UF - 30 Abril 2026", "uf_abril", previred_data)
    add_row("Tope AFP (90 UF)", "tope_afp_90uf", previred_data)
    add_row("Tope IPS (60 UF)", "tope_ips_60uf", previred_data)
    add_row("Tope Seg. Cesantía (135,2 UF)", "tope_cesantia_135uf", previred_data)
    add_row("Renta Mínima Dependientes", "renta_minima_dependientes", previred_data)
    add_row("Renta Mínima <18 / >65", "renta_minima_menores65", previred_data)
    add_row("Renta Mínima Casa Particular", "renta_minima_casa_particular", previred_data)
    add_row("Renta Mínima No Remuneracional", "renta_minima_no_remuneracional", previred_data)
    add_row("APV Tope Mensual (50 UF)", "apv_tope_mensual", previred_data)
    add_row("APV Tope Anual (600 UF)", "apv_tope_anual", previred_data)
    add_row("Depósito Convenido Tope Anual (900 UF)", "deposito_convenido_tope_anual", previred_data)
    add_row("Seguro Social", "seguro_social", previred_data, '%')
    add_row("Tasa SIS", "sis_tasa", previred_data, '%')
    add_row("Salud CCAF", "salud_ccaf", previred_data, '%')
    add_row("Salud FONASA", "salud_fonasa", previred_data, '%')
    write_sheet(ws, "Resumen", resumen_headers, resumen_rows)

    # Sheet 2: AFP
    if "AFP" in wb.sheetnames:
        ws_afp = wb["AFP"]
    else:
        ws_afp = wb.create_sheet("AFP")
    afp_headers = ["AFP", "Trabajador (%)", "Empleador (%)", "Total (%)", "Independiente (%)"]
    afp_names = ['Capital', 'Cuprum', 'Habitat', 'PlanVital', 'ProVida', 'Modelo', 'Uno']
    afp_rows = []
    for n in afp_names:
        k = n.lower()
        afp_rows.append([
            n,
            previred_data.get(f'afp_{k}_trabajador'),
            previred_data.get(f'afp_{k}_empleador'),
            previred_data.get(f'afp_{k}_total'),
            previred_data.get(f'afp_{k}_independiente')
        ])
    write_sheet(ws_afp, "AFP", afp_headers, afp_rows)

    # Sheet 3: AFC
    if "AFC" in wb.sheetnames:
        ws_afc = wb["AFC"]
    else:
        ws_afc = wb.create_sheet("AFC")
    afc_headers = ["Contrato", "Empleador (%)", "Trabajador (%)"]
    afc_rows = [
        ["Plazo Indefinido",
         previred_data.get('afc_indefinido_empleador'),
         previred_data.get('afc_indefinido_trabajador')],
        ["Plazo Fijo",
         previred_data.get('afc_plazofijo_empleador'), None],
        ["Indefinido 11+ años",
         previred_data.get('afc_11anios_empleador'), None],
        ["Casa Particular",
         previred_data.get('afc_casaparticular_empleador'), None],
    ]
    write_sheet(ws_afc, "AFC", afc_headers, afc_rows)

    # Sheet 4: Asignación Familiar
    if "AsignacionFamiliar" in wb.sheetnames:
        ws_afam = wb["AsignacionFamiliar"]
    else:
        ws_afam = wb.create_sheet("AsignacionFamiliar")
    afam_headers = ["Tramo", "Monto", "Requisito Renta"]
    afam_rows = [
        ["Tramo 1 (A)", previred_data.get('asignacion_familiar_tramo1'),
         previred_data.get('asignacion_familiar_tramo1_renta', '')],
        ["Tramo 2 (B)", previred_data.get('asignacion_familiar_tramo2'),
         previred_data.get('asignacion_familiar_tramo2_renta', '')],
        ["Tramo 3 (C)", previred_data.get('asignacion_familiar_tramo3'),
         previred_data.get('asignacion_familiar_tramo3_renta', '')],
        ["Tramo 4 (D)", previred_data.get('asignacion_familiar_tramo4_monto'),
         previred_data.get('asignacion_familiar_tramo4_renta', '')],
    ]
    write_sheet(ws_afam, "AsignacionFamiliar", afam_headers, afam_rows)

    # Sheet 5: Trabajos Pesados
    if "TrabajosPesados" in wb.sheetnames:
        ws_tp = wb["TrabajosPesados"]
    else:
        ws_tp = wb.create_sheet("TrabajosPesados")
    tp_headers = ["Tipo", "Empleador (%)", "Trabajador (%)"]
    tp_rows = [
        ["Trabajo Pesado", previred_data.get('trabajo_pesado_empleador'),
         previred_data.get('trabajo_pesado_trabajador')],
        ["Trabajo Menos Pesado", previred_data.get('trabajo_menos_pesado_empleador'),
         previred_data.get('trabajo_menos_pesado_trabajador')],
    ]
    write_sheet(ws_tp, "TrabajosPesados", tp_headers, tp_rows)

    # Sheet 6: UTM/UTA/IPC (SII mensual)
    if "UTM_UTA_IPC" in wb.sheetnames:
        ws_sii = wb["UTM_UTA_IPC"]
    else:
        ws_sii = wb.create_sheet("UTM_UTA_IPC")
    sii_headers = ["Mes", "UTM ($)", "UTA ($)", "IPC (puntos)", "Var. Mensual (%)", "Var. Acum. (%)", "Var. Anual (%)"]
    sii_rows = []
    for entry in sii_monthly:
        sii_rows.append([
            entry['mes'],
            entry['utm'], entry['uta'], entry['ipc'],
            entry['var_mensual'], entry['var_acum'], entry['var_anual']
        ])
    write_sheet(ws_sii, "UTM_UTA_IPC", sii_headers, sii_rows)

    # Sheet 7: Graficos
    add_charts_sheet(wb, sii_monthly)

    # Remove default sheet if exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(EXCEL_FILE)
    print(f"Excel guardado: {EXCEL_FILE}")

def save_to_json(previred_data, sii_data, sii_monthly):
    payload = {
        "ultima_actualizacion": datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        "año": 2026,
        "previred": {k: v for k, v in previred_data.items() if k != 'ultima_actualizacion'},
        "sii": {
            "mensual": sii_monthly,
            "fecha_consulta": sii_data.get('sii_fecha_consulta')
        }
    }
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"JSON guardado: {JSON_FILE}")

def main():
    print("=" * 50)
    print("Scraping Indicadores Previsionales - Previred")
    print("=" * 50)
    try:
        previred_data = scrape_previred()
        print(f"Previred OK - {len(previred_data)} campos extraídos")
    except Exception as e:
        print(f"Error scraping Previred: {e}")
        previred_data = {"error": str(e)}

    print("\n" + "=" * 50)
    print("Scraping UTM/UTA/IPC - SII")
    print("=" * 50)
    try:
        sii_data = scrape_sii_utm()
        sii_monthly = build_sii_monthly_list(sii_data)
        print(f"SII OK - {len(sii_monthly)} meses procesados")
    except Exception as e:
        print(f"Error scraping SII: {e}")
        sii_data = {"error": str(e)}
        sii_monthly = []

    print("\n" + "=" * 50)
    print("Guardando datos...")
    print("=" * 50)
    save_to_excel(previred_data, sii_data, sii_monthly)
    save_to_json(previred_data, sii_data, sii_monthly)

    print("\n¡Scraping completado!")

if __name__ == '__main__':
    main()
